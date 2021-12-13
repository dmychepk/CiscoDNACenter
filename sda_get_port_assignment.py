import dnac_credentials
import urllib3
import requests
from requests.auth import HTTPBasicAuth
from ratelimit import limits, sleep_and_retry
import logging
import json
from openpyxl import load_workbook


class Interface:
    def __init__(self, parameters):
        for k, v in parameters.items():
            setattr(self, k, v)


class Device:
    def __init__(self, hostname, device_type, ip, serial):
        self.hostname = hostname
        self.device_type = device_type
        self.ip = ip
        self.serial = serial


urllib3.disable_warnings()

logging.basicConfig(filename='port_assignment.log', format = '%(threadName)s: %(levelname)s: %(message)s', level=logging.INFO)

headers = {'Content-Type': 'application/json'}


def authenticate():
    auth_url = dnac_credentials.url + '/dna/system/api/v1/auth/token'
    response = requests.post(url=auth_url, auth=HTTPBasicAuth(dnac_credentials.username, dnac_credentials.password),
                             headers=headers, verify=False)
    headers['x-auth-token'] = response.json()['Token']


@sleep_and_retry
@limits(calls=100, period=60)
def get_port_assignment(device, interface):
    response = requests.get(
        url=dnac_credentials.url + f'/dna/intent/api/v1/business/sda/hostonboarding/user-device?'
        f'deviceManagementIpAddress={device}&'
        f'interfaceName={interface}',
        headers=headers, verify=False)
    if response.status_code != 200:
        raise Exception('API response: {}'.format(response.status_code))
    else:
        logging.info(f'{device} :: {interface} :: {response.json()}')
    return response


def get_device_assignment(device, fname):
    global count
    authenticate()
    intfs = []

    for j in range(0, len(device.device_type.split(','))):
        j += 1

        if device.device_type == 'Cisco Catalyst 9300 Switch':
            for i in range(0,36):
                i += 1
                interface = f'TwoGigabitEthernet{j}/0/{i}'
                api_response = get_port_assignment(device.ip, interface).json()
                if api_response.get('status') == 'success':
                    count += 1
                    intfs.append(Interface(api_response))
                    json.dump(api_response, fname, indent=4)
                    fname.write('\n')
            for i in range(36,48):
                i += 1
                interface = f'TenGigabitEthernet{j}/0/{i}'
                api_response = get_port_assignment(device.ip, interface).json()
                if api_response.get('status') == 'success':
                    count += 1
                    intfs.append(Interface(api_response))
                    json.dump(api_response, fname, indent=4)
                    fname.write('\n')

        elif device.device_type == 'Cisco Catalyst 35xx Stack-able Ethernet Switch':
            for i in range(0,6):
                i += 1
                interface = f'GigabitEthernet{j}/0/{i}'
                api_response = get_port_assignment(device.ip, interface).json()
                if api_response.get('status') == 'success':
                    count += 1
                    intfs.append(Interface(api_response))
                    json.dump(api_response, fname, indent=4)
                    fname.write('\n')
        elif device.device_type == 'Cisco Catalyst 9500 Switch':
            for i in range(0,40):
                i += 1
                interface = f'TenGigabitEthernet{j}/0/{i}'
                api_response = get_port_assignment(device.ip, interface).json()
                if api_response.get('status') == 'success':
                    count += 1
                    intfs.append(Interface(api_response))
                    json.dump(api_response, fname, indent=4)
                    fname.write('\n')

    return intfs


if __name__ == '__main__':

    devices = []
    wb = load_workbook('Switch Inventory.xlsx')
    ws = wb.active
    for row in ws.iter_rows(min_row=10, max_col=7, max_row=57, values_only=True):
        # (hostname, device_type, ip, serial)
        devices.append(Device(row[2], row[1], row[4], row[3]))

    count = 0

    interfaces = []
    f = open('port_assignment.json', 'w')

    for device in devices:
        interfaces.extend(get_device_assignment(device, f))

    for interface in interfaces:
        print(interface.__dict__)
    print(f'TOTAL: {count}')

    f.write(f'TOTAL: {count}')
    f.close()